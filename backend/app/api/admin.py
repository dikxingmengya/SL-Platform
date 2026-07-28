"""
管理端接口路由 /api/admin/*
管理员权限：Depends(RoleChecker(['admin']))
"""
from io import BytesIO

import openpyxl
from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from sqlalchemy import func, select, delete as sqla_delete
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.database import get_db
from app.models.course_type import CourseType
from app.models.lesson_record import LessonRecord
from app.models.student import Student
from app.models.user import User
from app.schemas.course_type import CourseTypeCreate, CourseTypeUpdate
from app.schemas.lesson_record import LessonRecordCreate
from app.schemas.package import PackageCreate, PackageUpdate
from app.schemas.student import StudentCreate, StudentUpdate
from app.schemas.teacher_student import TeacherStudentCreate
from app.schemas.user import UserCreate, UserUpdate
from app.services import (
    course_type_service,
    package_service,
    record_service,
    statistics_service,
    student_service,
    teacher_student_service,
    user_service,
)
from app.utils.permissions import RoleChecker, get_current_user
from app.utils.response import error, paginated_response, success
from app.utils.security import verify_password

router = APIRouter(prefix="/api/admin", tags=["管理端"])

# 所有管理端接口需要 admin 角色（路由级别依赖）
_admin_check = Depends(RoleChecker(["admin"]))


# ==================== 用户管理 ====================

@router.get("/users", summary="用户列表", dependencies=[_admin_check])
async def list_users(
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=20, ge=1, le=500),
    role: str | None = Query(default=None, description="角色筛选: admin/teacher/parent"),
    db: AsyncSession = Depends(get_db),
):
    """获取用户列表，支持按角色过滤"""
    users, total = await user_service.get_user_list(
        db, page=page, page_size=page_size, role=role
    )
    return paginated_response(items=users, total=total, page=page, page_size=page_size)


@router.post("/users", summary="创建用户", dependencies=[_admin_check])
async def create_user(
    data: UserCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """创建用户；创建管理员账号需超级管理员权限"""
    if data.role == "admin" and not current_user.is_super_admin:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="仅超级管理员可创建管理员账号")
    try:
        user = await user_service.create_user(db, data)
        return success(data={"id": user.id, "username": user.username}, msg="用户创建成功")
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))


@router.put("/users/{user_id}", summary="编辑用户", dependencies=[_admin_check])
async def update_user(
    user_id: int,
    data: UserUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """更新用户信息；修改密码需管理员身份验证；超级管理员不可被修改；仅超级管理员可变更管理员角色"""
    # 读取目标用户以判断权限
    target = await db.get(User, user_id)
    if not target:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="用户不存在")

    # 超级管理员账号受保护，不可被修改
    if target.is_super_admin:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="超级管理员账号不可修改")

    # 将普通用户提升为管理员、或修改管理员信息，需超级管理员权限
    is_admin_target = target.role == "admin" or (data.role and data.role == "admin")
    if is_admin_target and not current_user.is_super_admin:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="仅超级管理员可管理管理员账号")

    # 修改密码时二次验证管理员密码
    if data.password:
        if not data.admin_password:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="修改密码需要提供管理员密码验证")
        if not verify_password(data.admin_password, current_user.password_hash):
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="管理员密码验证失败")

    user = await user_service.update_user(db, user_id, data)
    return success(data={"id": user.id}, msg="用户更新成功")


@router.delete("/users/{user_id}", summary="删除用户", dependencies=[_admin_check])
async def delete_user(
    user_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """软删除用户；超级管理员不可删除"""
    target = await db.get(User, user_id)
    if not target:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="用户不存在")
    if target.is_super_admin:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="超级管理员账号不可删除")

    # 检查关联，无关联硬删除，有关联列出详情并阻止
    from app.models.teacher_student import TeacherStudent
    refs = []
    if target.role == "teacher":
        ts_rows = (await db.execute(select(TeacherStudent).where(TeacherStudent.teacher_id == user_id))).scalars().all()
        if ts_rows:
            names = [r.student.name for r in ts_rows if r.student]
            refs.append(f"{len(ts_rows)}个师生分配（{', '.join(names)}）")
        lr_count = (await db.execute(select(func.count(LessonRecord.id)).where(LessonRecord.teacher_id == user_id))).scalar() or 0
        if lr_count: refs.append(f"{lr_count}条上课记录")
    if target.role == "parent":
        sc_rows = (await db.execute(select(Student).where(Student.parent_user_id == user_id))).scalars().all()
        if sc_rows:
            names = [s.name for s in sc_rows]
            refs.append(f"{len(sc_rows)}个孩子（{', '.join(names)}）")
        pk = (await db.execute(select(func.count(Package.id)).where(Package.parent_user_id == user_id))).scalar() or 0
        if pk: refs.append(f"{pk}个课时包")
    if refs:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"禁止删除：{'；'.join(refs)}")
    await db.execute(sqla_delete(User).where(User.id == user_id))
    await db.commit()
    return success(msg="已删除")


# ==================== 学生档案管理 ====================

@router.get("/students", summary="学生列表", dependencies=[_admin_check])
async def list_students(
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=20, ge=1, le=500),
    db: AsyncSession = Depends(get_db),
):
    """获取学生列表（含所属家长姓名和电话）"""
    students, total = await student_service.get_student_list(
        db, page=page, page_size=page_size
    )
    return paginated_response(items=students, total=total, page=page, page_size=page_size)


@router.post("/students", summary="创建学生", dependencies=[_admin_check])
async def create_student(
    data: StudentCreate,
    db: AsyncSession = Depends(get_db),
):
    """创建学生档案，必须指定 parent_user_id（该校用户角色必须为 parent）"""
    try:
        student = await student_service.create_student(db, data)
        return success(
            data={"id": student.id, "name": student.name}, msg="学生档案创建成功"
        )
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))


@router.put("/students/{student_id}", summary="编辑学生", dependencies=[_admin_check])
async def update_student(
    student_id: int,
    data: StudentUpdate,
    db: AsyncSession = Depends(get_db),
):
    """编辑学生档案"""
    student = await student_service.update_student(db, student_id, data)
    if not student:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="学生不存在")
    return success(data={"id": student.id}, msg="学生档案更新成功")


@router.delete("/students/{student_id}", summary="删除学生", dependencies=[_admin_check])
async def delete_student(
    student_id: int,
    db: AsyncSession = Depends(get_db),
):
    """删除学生：无关联硬删除，有关联列出详情并阻止"""
    from app.models.teacher_student import TeacherStudent
    ts_rows = (await db.execute(select(TeacherStudent).where(TeacherStudent.student_id == student_id))).scalars().all()
    lr_count = (await db.execute(select(func.count(LessonRecord.id)).where(LessonRecord.student_id == student_id))).scalar() or 0
    refs = []
    if ts_rows:
        names = [r.teacher.real_name for r in ts_rows if r.teacher]
        refs.append(f"{len(ts_rows)}个师生分配（教师：{', '.join(names)}）")
    if lr_count: refs.append(f"{lr_count}条上课记录")
    if refs:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"禁止删除：{'；'.join(refs)}")
    await db.execute(sqla_delete(Student).where(Student.id == student_id))
    await db.commit()
    return success(msg="已删除")


# ==================== 师生分配管理 ====================

@router.get("/teacher-students", summary="师生分配列表", dependencies=[_admin_check])
async def list_assignments(
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=20, ge=1, le=500),
    teacher_id: int | None = Query(default=None),
    db: AsyncSession = Depends(get_db),
):
    """获取师生分配列表"""
    assignments, total = await teacher_student_service.get_assignments(
        db, page=page, page_size=page_size, teacher_id=teacher_id
    )
    return paginated_response(items=assignments, total=total, page=page, page_size=page_size)


@router.post("/teacher-students", summary="分配师生", dependencies=[_admin_check])
async def create_assignment(
    data: TeacherStudentCreate,
    db: AsyncSession = Depends(get_db),
):
    """将学生分配给教师"""
    try:
        assignment = await teacher_student_service.create_assignment(db, data)
        return success(data={"id": assignment.id}, msg="师生分配成功")
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))


@router.delete("/teacher-students/{assignment_id}", summary="取消分配", dependencies=[_admin_check])
async def delete_assignment(
    assignment_id: int,
    db: AsyncSession = Depends(get_db),
):
    """取消师生分配关系"""
    ok = await teacher_student_service.delete_assignment(db, assignment_id)
    if not ok:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="分配关系不存在")
    return success(msg="已取消分配")


# ==================== 课程类型管理 ====================

@router.get("/course-types", summary="课程类型列表", dependencies=[_admin_check])
async def list_course_types(
    db: AsyncSession = Depends(get_db),
):
    """获取所有课程类型"""
    types = await course_type_service.get_course_types(db)
    ct_list = [{
        "id": t.id,
        "name": t.name,
        "description": t.description,
        "default_hourly_rate": float(t.default_hourly_rate),
        "is_active": t.is_active,
        "created_at": t.created_at,
    } for t in types]
    return success(data=ct_list)


@router.post("/course-types", summary="新增课程类型", dependencies=[_admin_check])
async def create_course_type(
    data: CourseTypeCreate,
    db: AsyncSession = Depends(get_db),
):
    """新增课程类型"""
    ct = await course_type_service.create_course_type(db, data)
    return success(data={"id": ct.id, "name": ct.name}, msg="课程类型创建成功")


@router.put("/course-types/{ct_id}", summary="编辑课程类型", dependencies=[_admin_check])
async def update_course_type(
    ct_id: int,
    data: CourseTypeUpdate,
    db: AsyncSession = Depends(get_db),
):
    """编辑课程类型"""
    ct = await course_type_service.update_course_type(db, ct_id, data)
    if not ct:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="课程类型不存在")
    return success(data={"id": ct.id}, msg="课程类型更新成功")


@router.delete("/course-types/{ct_id}", summary="删除课程类型", dependencies=[_admin_check])
async def delete_course_type(
    ct_id: int,
    db: AsyncSession = Depends(get_db),
):
    """删除课程类型：有引用则阻止并告知引用位置"""
    from app.models.package import Package
    pk_count = (await db.execute(select(func.count(Package.id)).where(Package.course_type_id == ct_id))).scalar() or 0
    lr_count = (await db.execute(select(func.count(LessonRecord.id)).where(LessonRecord.course_type_id == ct_id))).scalar() or 0
    refs = []
    if pk_count:
        pk_rows = (await db.execute(select(Package).where(Package.course_type_id == ct_id).limit(5))).scalars().all()
        names = [f"ID#{p.id}" for p in pk_rows]
        refs.append(f"{pk_count}个课时包（{'、'.join(names)}{'等' if pk_count > 5 else ''}）")
    if lr_count: refs.append(f"{lr_count}条上课记录")
    if refs:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"禁止删除：{'；'.join(refs)}")
    await db.execute(sqla_delete(CourseType).where(CourseType.id == ct_id))
    await db.commit()
    return success(msg="已删除")
    return success(msg="已删除")


# ==================== 课时包管理 ====================

@router.get("/packages", summary="课时包列表", dependencies=[_admin_check])
async def list_packages(
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=20, ge=1, le=500),
    parent_user_id: int | None = Query(default=None),
    db: AsyncSession = Depends(get_db),
):
    """获取课时包列表"""
    packages, total = await package_service.get_packages(
        db, page=page, page_size=page_size, parent_user_id=parent_user_id
    )
    return paginated_response(items=packages, total=total, page=page, page_size=page_size)


@router.post("/packages", summary="购买课时包", dependencies=[_admin_check])
async def create_package(
    data: PackageCreate,
    db: AsyncSession = Depends(get_db),
):
    """为家长购买课时包（名下孩子共享）"""
    try:
        package = await package_service.create_package(db, data)
        return success(
            data={"id": package.id, "total_hours": float(package.total_hours)},
            msg="课时包购买成功",
        )
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))


@router.put("/packages/{package_id}", summary="手动调整课时包", dependencies=[_admin_check])
async def update_package(
    package_id: int,
    data: PackageUpdate,
    db: AsyncSession = Depends(get_db),
):
    """手动调整课时包信息"""
    package = await package_service.update_package(db, package_id, data)
    if not package:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="课时包不存在")
    return success(data={"id": package.id}, msg="课时包更新成功")


@router.delete("/packages/{package_id}", summary="删除课时包", dependencies=[_admin_check])
async def delete_package(
    package_id: int,
    db: AsyncSession = Depends(get_db),
):
    """删除课时包"""
    # 检查是否已消耗课时
    from app.models.package import Package
    pkg = await db.get(Package, package_id)
    if not pkg:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="课时包不存在")
    if float(pkg.used_hours) > 0:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"该课时包已消耗 {pkg.used_hours} 课时，无法删除。如需删除请先手动调整已消耗为0"
        )
    ok = await package_service.delete_package(db, package_id)
    return success(msg="已删除")


# ==================== 上课记录管理 ====================

@router.get("/records", summary="上课记录列表", dependencies=[_admin_check])
async def list_records(
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=20, ge=1, le=500),
    status: str | None = Query(default=None),
    teacher_id: int | None = Query(default=None),
    student_id: int | None = Query(default=None),
    parent_user_id: int | None = Query(default=None),
    reviewer_id: int | None = Query(default=None),
    db: AsyncSession = Depends(get_db),
):
    """获取所有上课记录（支持多条件筛选）"""
    records, total = await record_service.get_records(
        db, page=page, page_size=page_size, status=status,
        teacher_id=teacher_id, student_id=student_id,
        parent_user_id=parent_user_id, reviewer_id=reviewer_id,
    )
    return paginated_response(items=records, total=total, page=page, page_size=page_size)


@router.put("/records/{record_id}", summary="编辑上课记录", dependencies=[_admin_check])
async def update_record(
    record_id: int,
    data: LessonRecordCreate,
    db: AsyncSession = Depends(get_db),
):
    """管理员编辑上课记录"""
    result = await db.execute(
        select(LessonRecord)
        .options(selectinload(LessonRecord.student), selectinload(LessonRecord.teacher),
                 selectinload(LessonRecord.course_type))
        .where(LessonRecord.id == record_id)
    )
    record = result.scalar_one_or_none()
    if not record:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="记录不存在")
    record.student_id = data.student_id
    record.course_type_id = data.course_type_id
    record.hours = data.hours
    record.date = data.date
    record.content = data.content
    await db.commit()
    return success(data={"id": record.id}, msg="记录已更新")


@router.delete("/records/{record_id}", summary="删除上课记录", dependencies=[_admin_check])
async def delete_record(
    record_id: int,
    db: AsyncSession = Depends(get_db),
):
    """
    删除上课记录
    若状态为 approved，则退回已扣减的课时（按 expire_date DESC，后扣的先退）
    """
    from app.models.package import Package

    result = await db.execute(
        select(LessonRecord)
        .options(selectinload(LessonRecord.student))
        .where(LessonRecord.id == record_id)
    )
    record = result.scalar_one_or_none()
    if not record:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="记录不存在")

    # 已审核通过的记录，需退回课时
    if record.status == "approved" and record.student:
        parent_id = record.student.parent_user_id
        refund_remaining = float(record.hours)

        # 按 expire_date DESC 退回（后扣的先退，LIFO）
        packages = (await db.execute(
            select(Package)
            .where(
                Package.parent_user_id == parent_id,
                (Package.course_type_id == record.course_type_id) | (Package.course_type_id.is_(None)),
                Package.used_hours > 0,
            )
            .order_by(Package.expire_date.desc())
            .with_for_update()
        )).scalars().all()

        for pkg in packages:
            refund = min(float(pkg.used_hours), refund_remaining)
            pkg.used_hours = float(pkg.used_hours) - refund
            refund_remaining -= refund
            # 恢复状态
            if pkg.status == "depleted" and float(pkg.used_hours) < float(pkg.total_hours):
                pkg.status = "active"
            if refund_remaining <= 0:
                break

    await db.delete(record)
    await db.commit()
    msg = "已删除，课时已退回" if record.status == "approved" else "已删除"
    return success(msg=msg)


@router.put("/records/{record_id}/approve", summary="通过审核", dependencies=[_admin_check])
async def approve_record(
    record_id: int,
    comment: str = Query(default="", description="审核意见"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """审核通过 + 原子扣减课时"""
    try:
        record = await record_service.approve_record(
            db, record_id=record_id, reviewer_id=current_user.id, comment=comment
        )
        return success(
            data={"id": record.id, "status": record.status, "hours": float(record.hours)},
            msg="审核通过，课时已自动扣减",
        )
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))


@router.put("/records/{record_id}/reject", summary="驳回记录", dependencies=[_admin_check])
async def reject_record(
    record_id: int,
    comment: str = Query(default="", description="驳回原因"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """驳回上课记录"""
    try:
        record = await record_service.reject_record(
            db, record_id=record_id, reviewer_id=current_user.id, comment=comment
        )
        if not record:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="记录不存在")
        return success(data={"id": record.id, "status": record.status}, msg="记录已驳回")
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))


@router.get("/records/export", summary="导出上课记录Excel", dependencies=[_admin_check])
async def export_records(
    status: str | None = Query(default=None),
    teacher_id: int | None = Query(default=None),
    student_id: int | None = Query(default=None),
    parent_user_id: int | None = Query(default=None),
    reviewer_id: int | None = Query(default=None),
    db: AsyncSession = Depends(get_db),
):
    """导出上课记录为Excel（支持按筛选条件导出）"""
    records, _ = await record_service.get_records(
        db, page=1, page_size=10000, status=status,
        teacher_id=teacher_id, student_id=student_id,
        parent_user_id=parent_user_id, reviewer_id=reviewer_id,
    )
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "上课记录"
    headers = ["ID","学生","教师","课程","课时","上课时间","内容","状态","审核人","审核意见","审核时间"]
    for c, h in enumerate(headers, 1): ws.cell(row=1, column=c, value=h)
    for r, rec in enumerate(records, 2):
        for c, v in enumerate([rec["id"],rec["student_name"],rec["teacher_name"],rec["course_type_name"],
            rec["hours"],str(rec["date"]),rec["content"],rec["status"],rec["reviewer_name"],
            rec["review_comment"],str(rec["reviewed_at"] or "")], 1):
            ws.cell(row=r, column=c, value=v)
    out = BytesIO(); wb.save(out); out.seek(0)
    return StreamingResponse(out, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=records.xlsx"})


# ==================== 统计报表 ====================

@router.get("/statistics/overview", summary="统计概览", dependencies=[_admin_check])
async def get_overview(
    db: AsyncSession = Depends(get_db),
):
    """获取系统概览统计数据"""
    overview = await statistics_service.get_overview_statistics(db)
    monthly = await statistics_service.get_monthly_trend(db)
    revenue = await statistics_service.get_revenue_trend(db)
    return success(data={
        "overview": overview,
        "monthly_trend": monthly,
        "revenue_trend": revenue,
    })


@router.get("/statistics/export", summary="导出Excel报表", dependencies=[_admin_check])
async def export_statistics(
    db: AsyncSession = Depends(get_db),
):
    """导出统计报表为 Excel 文件下载"""
    excel_bytes = await statistics_service.export_statistics_excel(db)
    return StreamingResponse(
        excel_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=sl-platform-statistics.xlsx"},
    )
