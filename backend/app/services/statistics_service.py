"""
统计服务：数据概览、按月趋势、Excel 导出
"""
from datetime import date, datetime
from io import BytesIO
from typing import Any, Optional

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy import func, select, text
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.course_type import CourseType
from app.models.lesson_record import LessonRecord
from app.models.package import Package
from app.models.student import Student
from app.models.user import User


async def get_overview_statistics(db: AsyncSession) -> dict:
    """
    获取系统概览统计数据
    """
    # 各类用户数量
    total_students = (await db.execute(
        select(func.count(Student.id)).where(Student.is_active == True)
    )).scalar() or 0

    total_teachers = (await db.execute(
        select(func.count(User.id)).where(
            User.role == "teacher", User.is_active == True
        )
    )).scalar() or 0

    total_parents = (await db.execute(
        select(func.count(User.id)).where(
            User.role == "parent", User.is_active == True
        )
    )).scalar() or 0

    # 课时包统计
    total_packages = (await db.execute(
        select(func.count(Package.id))
    )).scalar() or 0

    total_package_hours = (await db.execute(
        select(func.coalesce(func.sum(Package.total_hours), 0))
    )).scalar() or 0

    total_used_hours = (await db.execute(
        select(func.coalesce(func.sum(Package.used_hours), 0))
    )).scalar() or 0

    total_package_revenue = (await db.execute(
        select(func.coalesce(func.sum(Package.price), 0))
    )).scalar() or 0

    # 上课记录统计
    total_records = (await db.execute(
        select(func.count(LessonRecord.id))
    )).scalar() or 0

    total_approved_hours = (await db.execute(
        select(func.coalesce(func.sum(LessonRecord.hours), 0)).where(
            LessonRecord.status == "approved"
        )
    )).scalar() or 0

    pending_records = (await db.execute(
        select(func.count(LessonRecord.id)).where(
            LessonRecord.status == "pending"
        )
    )).scalar() or 0

    return {
        "total_students": total_students,
        "total_teachers": total_teachers,
        "total_parents": total_parents,
        "total_packages": total_packages,
        "total_package_hours": float(total_package_hours),
        "total_used_hours": float(total_used_hours),
        "remaining_hours": float(total_package_hours) - float(total_used_hours),
        "total_package_revenue": float(total_package_revenue),
        "total_records": total_records,
        "total_approved_hours": float(total_approved_hours),
        "pending_records": pending_records,
    }


async def get_monthly_trend(db: AsyncSession) -> list[dict]:
    """
    按月统计上课趋势（SQL 聚合函数）
    使用 GROUP BY DATE_FORMAT 按月分组
    """
    # 使用原生 SQL 进行按月统计，因为 SQLAlchemy 的 DATE_FORMAT 支持有限
    result = await db.execute(
        text(
            """
            SELECT
                DATE_FORMAT(lr.date, '%Y-%m') AS month,
                COUNT(*) AS lesson_count,
                COALESCE(SUM(lr.hours), 0) AS total_hours,
                COUNT(DISTINCT lr.teacher_id) AS teacher_count,
                COUNT(DISTINCT lr.student_id) AS student_count
            FROM lesson_record lr
            WHERE lr.status = 'approved'
            GROUP BY DATE_FORMAT(lr.date, '%Y-%m')
            ORDER BY month ASC
            """
        )
    )
    rows = result.fetchall()

    return [
        {
            "month": row[0],
            "lesson_count": row[1],
            "total_hours": float(row[2]),
            "teacher_count": row[3],
            "student_count": row[4],
        }
        for row in rows
    ]


async def get_revenue_trend(db: AsyncSession) -> list[dict]:
    """
    按月统计收入趋势
    """
    result = await db.execute(
        text(
            """
            SELECT
                DATE_FORMAT(p.created_at, '%Y-%m') AS month,
                COUNT(*) AS package_count,
                COALESCE(SUM(p.price), 0) AS total_revenue
            FROM package p
            GROUP BY DATE_FORMAT(p.created_at, '%Y-%m')
            ORDER BY month ASC
            """
        )
    )
    rows = result.fetchall()

    return [
        {
            "month": row[0],
            "package_count": row[1],
            "total_revenue": float(row[2]),
        }
        for row in rows
    ]


async def export_statistics_excel(db: AsyncSession) -> BytesIO:
    """
    导出统计报表为 Excel 文件
    """
    wb = openpyxl.Workbook()

    # 样式定义
    header_font = Font(name="微软雅黑", bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    cell_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def style_header(ws, row, cols):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

    def style_cell(ws, row, cols):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = cell_alignment
            cell.border = thin_border

    # ---- Sheet 1: 概览统计 ----
    ws1 = wb.active
    ws1.title = "概览统计"
    overview = await get_overview_statistics(db)

    headers1 = ["指标", "数值"]
    for col, h in enumerate(headers1, 1):
        ws1.cell(row=1, column=col, value=h)
    style_header(ws1, 1, 2)

    overview_data = [
        ("学生总数", overview["total_students"]),
        ("教师总数", overview["total_teachers"]),
        ("家长总数", overview["total_parents"]),
        ("课时包总数", overview["total_packages"]),
        ("课时包总课时", overview["total_package_hours"]),
        ("已消耗课时", overview["total_used_hours"]),
        ("剩余课时", overview["remaining_hours"]),
        ("课时包总收入(元)", overview["total_package_revenue"]),
        ("上课记录总数", overview["total_records"]),
        ("已审核通过课时", overview["total_approved_hours"]),
        ("待审核记录数", overview["pending_records"]),
    ]
    for row_idx, (label, value) in enumerate(overview_data, 2):
        ws1.cell(row=row_idx, column=1, value=label)
        ws1.cell(row=row_idx, column=2, value=value)
        style_cell(ws1, row_idx, 2)
    ws1.column_dimensions["A"].width = 22
    ws1.column_dimensions["B"].width = 15

    # ---- Sheet 2: 上课趋势 ----
    ws2 = wb.create_sheet("按月上课统计")
    monthly = await get_monthly_trend(db)

    headers2 = ["月份", "上课次数", "总课时", "授课教师数", "上课学生数"]
    for col, h in enumerate(headers2, 1):
        ws2.cell(row=1, column=col, value=h)
    style_header(ws2, 1, 5)

    for row_idx, item in enumerate(monthly, 2):
        ws2.cell(row=row_idx, column=1, value=item["month"])
        ws2.cell(row=row_idx, column=2, value=item["lesson_count"])
        ws2.cell(row=row_idx, column=3, value=item["total_hours"])
        ws2.cell(row=row_idx, column=4, value=item["teacher_count"])
        ws2.cell(row=row_idx, column=5, value=item["student_count"])
        style_cell(ws2, row_idx, 5)
    for col_letter in ["A", "B", "C", "D", "E"]:
        ws2.column_dimensions[col_letter].width = 16

    # ---- Sheet 3: 收入趋势 ----
    ws3 = wb.create_sheet("按月收入统计")
    revenue = await get_revenue_trend(db)

    headers3 = ["月份", "课时包数量", "总收入(元)"]
    for col, h in enumerate(headers3, 1):
        ws3.cell(row=1, column=col, value=h)
    style_header(ws3, 1, 3)

    for row_idx, item in enumerate(revenue, 2):
        ws3.cell(row=row_idx, column=1, value=item["month"])
        ws3.cell(row=row_idx, column=2, value=item["package_count"])
        ws3.cell(row=row_idx, column=3, value=item["total_revenue"])
        style_cell(ws3, row_idx, 3)
    for col_letter in ["A", "B", "C"]:
        ws3.column_dimensions[col_letter].width = 18

    # 保存到内存
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
